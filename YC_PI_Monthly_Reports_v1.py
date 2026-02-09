iimport re
import traceback
from io import BytesIO
import zipfile

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ============================================================
# Yellow Cluster: Budget Summary Generator
# - Supports selecting database type:
#     * PPM (Aggie Enterprise / PPM export)
#     * GL  (General Ledger export)
# - Award Info Document is OPTIONAL (PPM only)
# - GL:
#     * User chooses sheet
#     * Header row auto-detected WITHOUT needing exact header row
#     * Column names auto-detected (keywords)
#     * Ignores Category == "N/A" (and blanks)
#     * Pulls Description + Financial Department
#     * Adds Notes (blank) columns for manual entry
#     * Output workbook:
#         - GL_Subset: transactions (selected cols + Financial Department + Description + Notes)
#         - GL_Pulled_Columns_Check: ONLY pulled cols (no Notes) for verification
#         - Pivot_By_Category:
#           Category | Budget(blank) | Actuals | Balance (=Budget-Actuals) | Notes(blank)
#           + totals row with SUM formulas (always sums rows above)
#           + Balance conditional colors (green positive / red negative)
#           + Title: "<Financial Department> Expenses"
#         - Larger fonts + alternating light-green rows for readability
# - UI:
#     * If GL selected, Award uploader + indirect toggle hidden
# ============================================================

# -----------------------------
# PAGE CONFIG (must be first Streamlit call)
# -----------------------------
st.set_page_config(
    page_title="Yellow Cluster: Budget Summary Generator",
    page_icon="🐄",
    layout="wide",
)

# -----------------------------
# SETTINGS
# -----------------------------
SUMMARY_SHEET_NAME = "Summary"
HEADER_ROW_INDEX = 17  # 0-based index: Excel row 18 (i.e., delete first 17 rows)
REPORT_RUNDATE_ROW0 = 2
REPORT_RUNDATE_COL0 = 0

# PPM column identifiers
PI_COL_NAME = "Project Principal Investigator"
PROJECT_COL_NAME = "Project Number"
TASK_NAME_COL_NAME = "Task Name"
TASK_NUMBER_COL_NAME = "Task Number"
STATUS_COL_NAME = "Project Status"

ALLOC_BUDGET_NET_COL = "Allocated Budget*"
CURRENT_BAL_NET_COL = "Current Balance*"

# GL canonical columns
GL_COL_CATEGORY = "Category"
GL_COL_ACTUALS = "Actuals"
GL_COL_PERIOD = "Accounting Period"
GL_COL_ACTIVITY = "Activity Code"
GL_COL_FIN_DEPT = "Financial Department"
GL_COL_DESC = "Description"

# -----------------------------
# Helpers
# -----------------------------
def normalize_columns(cols):
    return [str(c).replace("\xa0", " ").strip() for c in cols]


def safe_str(x) -> str:
    if x is None:
        return ""
    if isinstance(x, float) and pd.isna(x):
        return ""
    return str(x).replace("\xa0", " ").strip()


def extract_report_run_date_from_cell(val) -> str:
    s = safe_str(val)
    if not s:
        return ""
    m = re.search(r"(\d{4}-\d{2}-\d{2})", s)
    return m.group(1) if m else ""


def canon_key(x) -> str:
    s = safe_str(x)
    if not s:
        return ""
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return re.sub(r"[^A-Za-z0-9]", "", s).upper()


def find_column_by_exact_or_keywords(columns, target_name, keywords=None):
    columns = list(columns)
    if target_name in columns:
        return target_name

    if keywords:
        lowered = [c.lower() for c in columns]
        for col, low in zip(columns, lowered):
            if all(k.lower() in low for k in keywords):
                return col

    raise KeyError(
        f"Could not find a suitable column for '{target_name}'. Available columns: {columns}"
    )


def normalize_pi_last_first(pi_val: str) -> str:
    s = safe_str(pi_val)
    if not s:
        return ""
    if "," in s:
        return s
    parts = s.split()
    if len(parts) >= 2:
        return f"{parts[-1]}, {' '.join(parts[:-1])}"
    return s


def make_safe_filename_fragment(name: str) -> str:
    frag = safe_str(name)
    frag = re.sub(r'[\/:*?"<>|]+', "_", frag)
    frag = frag.strip().strip(".")
    return frag[:120] if frag else "Report"


# -----------------------------
# Formatting helpers
# -----------------------------
def _auto_width(ws, max_width=55, min_width=10):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _style_table_sheet(ws, header_row=1, start_data_row=2, currency_headers=None, font_size=12):
    """
    - Bold header
    - Larger fonts
    - Alternating light-green row shading
    - Currency formatting for specified header names
    """
    currency_headers = currency_headers or []

    header_font = Font(bold=True, size=font_size + 1)
    body_font = Font(size=font_size)

    # Header styling
    for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row)):
        cell.font = header_font

    # Map headers to column letters
    headers = [c.value for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    header_to_col = {h: get_column_letter(i + 1) for i, h in enumerate(headers)}

    # Alternating fill
    fill_green = PatternFill(start_color="FFE6F4EA", end_color="FFE6F4EA", fill_type="solid")

    for r in range(start_data_row, ws.max_row + 1):
        for cell in ws[r]:
            cell.font = body_font
            if (r - start_data_row) % 2 == 1:
                cell.fill = fill_green

    # Currency formatting
    for h in currency_headers:
        if h in header_to_col:
            col_letter = header_to_col[h]
            for cell in ws[col_letter]:
                if cell.row <= header_row:
                    continue
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


# -----------------------------
# PPM readers
# -----------------------------
def read_aggy_master(master_bytes: bytes):
    df_raw = pd.read_excel(BytesIO(master_bytes), sheet_name=SUMMARY_SHEET_NAME, header=None)

    report_cell = None
    try:
        report_cell = df_raw.iat[REPORT_RUNDATE_ROW0, REPORT_RUNDATE_COL0]
    except Exception:
        report_cell = None
    report_date = extract_report_run_date_from_cell(report_cell)

    header = df_raw.iloc[HEADER_ROW_INDEX]
    df = df_raw.iloc[HEADER_ROW_INDEX + 1 :].copy()
    df.columns = header
    df = df.dropna(how="all")
    df.columns = normalize_columns(df.columns)
    return df, report_date


def read_award(award_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    xl = pd.ExcelFile(BytesIO(award_bytes))
    df = pd.read_excel(xl, sheet_name=sheet_name)
    df.columns = normalize_columns(df.columns)
    return df


# -----------------------------
# GL: header auto-detection
# -----------------------------
def normalize_header_cell(x) -> str:
    return safe_str(x).replace("\n", " ").strip()


def best_effort_match(col: str, required: str) -> bool:
    c = normalize_header_cell(col).lower()
    if not c:
        return False
    if c == required.lower():
        return True

    req_keywords = {
        GL_COL_CATEGORY: ["category"],
        GL_COL_ACTUALS: ["actual"],
        GL_COL_PERIOD: ["accounting", "period"],
        GL_COL_ACTIVITY: ["activity", "code"],
        GL_COL_FIN_DEPT: ["financial", "department"],
        GL_COL_DESC: ["description"],
    }
    kws = req_keywords.get(required, [required.lower()])
    return all(k in c for k in kws)


def detect_header_row_from_required_cols(raw_df: pd.DataFrame, required_cols: list, search_rows: int = 80) -> int:
    best_row = None
    best_score = -1

    max_r = min(search_rows, len(raw_df))
    for r in range(max_r):
        headers = [normalize_header_cell(v) for v in raw_df.iloc[r].tolist()]

        score = 0
        for req in required_cols:
            if any(best_effort_match(h, req) for h in headers):
                score += 1

        if score > best_score:
            best_score = score
            best_row = r

        if best_score == len(required_cols):
            break

    # Require at least len(required_cols)-1 matches (robust but not too strict)
    min_needed = max(1, len(required_cols) - 1)
    if best_row is None or best_score < min_needed:
        raise KeyError(
            f"Could not detect a header row with at least {min_needed} of {len(required_cols)} "
            f"required columns: {required_cols}. Tried first {max_r} rows."
        )

    return best_row


def read_gl_with_auto_header(db_bytes: bytes, sheet_name: str, required_cols: list[str], search_rows: int = 80):
    xl = pd.ExcelFile(BytesIO(db_bytes))
    raw = pd.read_excel(xl, sheet_name=sheet_name, header=None)
    header_row = detect_header_row_from_required_cols(raw, required_cols=required_cols, search_rows=search_rows)
    df = pd.read_excel(xl, sheet_name=sheet_name, header=header_row)
    df.columns = normalize_columns(df.columns)
    return df, header_row


# -----------------------------
# PPM output builder (ZIP per PI)
# -----------------------------
def build_pi_zip(df_out: pd.DataFrame, pi_col: str, hide_indirect: bool, report_label: str) -> bytes:
    df_out = df_out.copy()
    df_out[pi_col] = df_out[pi_col].apply(normalize_pi_last_first)

    unique_pis = [p for p in df_out[pi_col].dropna().unique().tolist() if safe_str(p)]
    unique_pis_sorted = sorted(unique_pis, key=lambda s: safe_str(s).lower())

    applied_indirects = False
    if "Indirect Rate" in df_out.columns:
        try:
            applied_indirects = pd.to_numeric(df_out["Indirect Rate"], errors="coerce").fillna(0).abs().gt(0).any()
        except Exception:
            applied_indirects = False

    footnote = (
        "* Calculated minus the indirect costs, if applicable."
        if applied_indirects
        else "* Indirect costs not applied (no Award Info document, or no indirect rates found)."
    )

    zip_buf = BytesIO()
    used_names = set()

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for pi in unique_pis_sorted:
            group = df_out[df_out[pi_col] == pi].copy()
            if group.empty:
                continue

            if ALLOC_BUDGET_NET_COL in group.columns:
                group[ALLOC_BUDGET_NET_COL] = pd.to_numeric(group[ALLOC_BUDGET_NET_COL], errors="coerce")
                group = group.sort_values(by=ALLOC_BUDGET_NET_COL, ascending=False, na_position="last")

            currency_cols = [c for c in [ALLOC_BUDGET_NET_COL, CURRENT_BAL_NET_COL, "expenses"] if c in group.columns]

            xbuf = BytesIO()
            with pd.ExcelWriter(xbuf, engine="openpyxl") as writer:
                group.to_excel(writer, index=False, sheet_name="Budget Summary")
                wb = writer.book

                # Style
                ws = wb["Budget Summary"]
                # Use existing style logic but bump font sizes and add alternating shading
                _style_table_sheet(ws, header_row=1, start_data_row=2, currency_headers=currency_cols, font_size=12)

                # Optional hide indirect column
                if hide_indirect and "Indirect Rate" in group.columns:
                    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    if "Indirect Rate" in headers:
                        idx = headers.index("Indirect Rate") + 1
                        ws.column_dimensions[get_column_letter(idx)].hidden = True

                _auto_width(ws)

                footer_row = ws.max_row + 2
                ws[f"A{footer_row}"] = footnote
                ws[f"A{footer_row}"].font = Font(italic=True, size=11)

            xbuf.seek(0)

            safe_pi = make_safe_filename_fragment(pi)
            filename = f"{report_label} - {safe_pi}.xlsx"

            if filename in used_names:
                k = 2
                while True:
                    candidate = f"{report_label} - {safe_pi} ({k}).xlsx"
                    if candidate not in used_names:
                        filename = candidate
                        break
                    k += 1
            used_names.add(filename)

            zf.writestr(filename, xbuf.read())

    zip_buf.seek(0)
    return zip_buf.getvalue()


# -----------------------------
# GL output builder (3-sheet workbook)
# -----------------------------
def build_gl_excel(df_subset: pd.DataFrame, pivot_categories_actuals: pd.DataFrame, dept_name: str) -> bytes:
    dept_name_clean = safe_str(dept_name) or "Financial Department"
    title = f"{dept_name_clean} Expenses"

    # Sheet: only pulled columns (verification)
    check_cols = [GL_COL_CATEGORY, GL_COL_ACTUALS, GL_COL_PERIOD, GL_COL_ACTIVITY, GL_COL_FIN_DEPT, GL_COL_DESC]
    df_check = df_subset[[c for c in check_cols if c in df_subset.columns]].copy()

    # Pivot base (Category + Actuals)
    piv = pivot_categories_actuals.copy()
    piv = piv.rename(columns={GL_COL_ACTUALS: "Actuals"})
    piv = piv.sort_values(by="Actuals", ascending=False, na_position="last")

    # Add blank/manual columns
    piv.insert(1, "Budget", "")
    piv["Balance"] = ""   # formula in Excel
    piv["Notes"] = ""     # manual

    xbuf = BytesIO()
    with pd.ExcelWriter(xbuf, engine="openpyxl") as writer:
        df_subset.to_excel(writer, index=False, sheet_name="GL_Subset")
        df_check.to_excel(writer, index=False, sheet_name="GL_Pulled_Columns_Check")
        piv.to_excel(writer, index=False, sheet_name="Pivot_By_Category")

        wb = writer.book

        # ---- GL_Subset styling ----
        ws_sub = wb["GL_Subset"]
        _style_table_sheet(ws_sub, header_row=1, start_data_row=2, currency_headers=[GL_COL_ACTUALS], font_size=12)
        _auto_width(ws_sub)

        # ---- Check sheet styling ----
        ws_chk = wb["GL_Pulled_Columns_Check"]
        _style_table_sheet(ws_chk, header_row=1, start_data_row=2, currency_headers=[GL_COL_ACTUALS], font_size=12)
        _auto_width(ws_chk)

        # ---- Pivot styling + formulas ----
        ws_piv = wb["Pivot_By_Category"]

        # Insert title row
        ws_piv.insert_rows(1)
        ws_piv["A1"] = title
        ws_piv["A1"].font = Font(bold=True, size=14)

        # Style pivot (header row now 2, data starts 3)
        _style_table_sheet(ws_piv, header_row=2, start_data_row=3, currency_headers=["Budget", "Actuals", "Balance"], font_size=12)

        headers = [c.value for c in next(ws_piv.iter_rows(min_row=2, max_row=2))]
        col_map = {name: idx + 1 for idx, name in enumerate(headers)}  # 1-based index

        data_start = 3
        data_end = ws_piv.max_row  # current last row (categories)

        # Add Balance formulas for all category rows (Balance = Budget - Actuals)
        b_col = col_map.get("Budget")
        a_col = col_map.get("Actuals")
        bal_col = col_map.get("Balance")

        if b_col and a_col and bal_col:
            for r in range(data_start, data_end + 1):
                b_cell = ws_piv.cell(row=r, column=b_col).coordinate
                a_cell = ws_piv.cell(row=r, column=a_col).coordinate
                ws_piv.cell(row=r, column=bal_col).value = f"={b_cell}-{a_cell}"

            # Conditional formatting on Balance (green positive / red negative)
            bal_letter = get_column_letter(bal_col)
            rng = f"{bal_letter}{data_start}:{bal_letter}{data_end}"
            ws_piv.conditional_formatting.add(
                rng,
                CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="004B00", bold=True)),
            )
            ws_piv.conditional_formatting.add(
                rng,
                CellIsRule(operator="lessThan", formula=["0"], font=Font(color="8B0000", bold=True)),
            )

        # Totals row with SUM formulas (always sums rows above)
        total_row = ws_piv.max_row + 1
        ws_piv.cell(row=total_row, column=col_map["Category"]).value = "TOTAL"

        def set_sum_total(col_name: str):
            if col_name not in col_map:
                return
            c = col_map[col_name]
            col_letter = get_column_letter(c)
            ws_piv.cell(row=total_row, column=c).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            ws_piv.cell(row=total_row, column=c).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        set_sum_total("Budget")
        set_sum_total("Actuals")
        set_sum_total("Balance")

        # Style totals row
        fill = PatternFill(start_color="FFFAD7", end_color="FFFAD7", fill_type="solid")
        for cell in ws_piv[total_row]:
            cell.font = Font(bold=True, size=13)
            cell.fill = fill

        _auto_width(ws_piv)

    xbuf.seek(0)
    return xbuf.getvalue()


# -----------------------------
# UI
# -----------------------------
st.markdown(
    """
    <div style="padding: 1rem 1.25rem; border-radius: 12px; background: #01223d; color: white; margin-bottom: 1rem;">
      <div style="font-size: 1.35rem; font-weight: 700;">🐄 Yellow Cluster • Budget Report Generator</div>
      <div style="opacity: 0.85; margin-top: 0.25rem;">
        Generate PI-level budget reports from a PPM (Aggie Enterprise) export or build a GL expenses workbook.
        <br/>Report bugs to David Railton Garrett drgarrett@ucdavis.edu
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# Best-effort radio blue (reliable method is Streamlit theme config)
st.markdown(
    """
    <style>
      div[role="radiogroup"] input[type="radio"]:checked + div {
        background: #1f77ff !important;
        border-color: #1f77ff !important;
      }
      div[role="radiogroup"] input[type="radio"] + div {
        border-color: #1f77ff !important;
      }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown("### Step 1 — Choose database type")
db_type = st.radio(
    "Which database are you uploading?",
    options=["PPM", "GL"],
    index=0,
    horizontal=True,
    help="PPM = Aggie Enterprise (PPM export). GL = General Ledger export.",
)
st.markdown("---")

with st.expander("Debug options", expanded=False):
    show_trace = st.checkbox("Show full error trace", value=False)
    show_unmatched = st.checkbox("Show unmatched key samples", value=True)
    show_key_samples = st.checkbox("Show key samples from both files", value=True)

db_file = st.file_uploader(
    "Upload PPM Database (Aggie Enterprise / PPM export)" if db_type == "PPM" else "Upload GL Database (General Ledger export)",
    type=["xlsx"],
)

award_file = None
hide_indirect_in_output = True
if db_type == "PPM":
    award_file = st.file_uploader("Upload Award Info Document (Excel) — optional", type=["xlsx"])
    hide_indirect_in_output = st.checkbox("Hide 'Indirect Rate' column in resulting download", value=True)

if db_file:
    try:
        db_bytes = db_file.getvalue()

        # ============================================================
        # GL MODE
        # ============================================================
        if db_type == "GL":
            xl_gl = pd.ExcelFile(BytesIO(db_bytes))
            gl_sheet = st.selectbox("GL sheet to use", options=xl_gl.sheet_names, index=0)

            # Header detection should rely on the 4 core columns (more robust),
            # NOT on Financial Department / Description (which may vary by export)
            required_for_header = [GL_COL_CATEGORY, GL_COL_ACTUALS, GL_COL_PERIOD, GL_COL_ACTIVITY]
            df_gl, detected_header_row0 = read_gl_with_auto_header(
                db_bytes=db_bytes,
                sheet_name=gl_sheet,
                required_cols=required_for_header,
                search_rows=80,
            )
            st.caption(f"Detected header row at Excel row **{detected_header_row0 + 1}**.")

            override = st.checkbox("Override detected header row", value=False)
            if override:
                override_row_excel = st.number_input(
                    "Header row (Excel row number)",
                    min_value=1,
                    value=int(detected_header_row0 + 1),
                    step=1,
                )
                df_gl = pd.read_excel(pd.ExcelFile(BytesIO(db_bytes)), sheet_name=gl_sheet, header=int(override_row_excel) - 1)
                df_gl.columns = normalize_columns(df_gl.columns)
                st.caption(f"Using overridden header row at Excel row **{int(override_row_excel)}**.")

            # Detect needed columns from the detected header
            col_category = find_column_by_exact_or_keywords(df_gl.columns, GL_COL_CATEGORY, keywords=["category"])
            col_actuals = find_column_by_exact_or_keywords(df_gl.columns, GL_COL_ACTUALS, keywords=["actual"])
            col_period = find_column_by_exact_or_keywords(df_gl.columns, GL_COL_PERIOD, keywords=["accounting", "period"])
            col_activity = find_column_by_exact_or_keywords(df_gl.columns, GL_COL_ACTIVITY, keywords=["activity", "code"])
            col_fin_dept = find_column_by_exact_or_keywords(df_gl.columns, GL_COL_FIN_DEPT, keywords=["financial", "department"])
            col_desc = find_column_by_exact_or_keywords(df_gl.columns, GL_COL_DESC, keywords=["description"])

            # Subset to selected columns (include Financial Department + Description)
            df_subset = df_gl[[col_category, col_actuals, col_period, col_activity, col_fin_dept, col_desc]].copy()
            df_subset = df_subset.rename(
                columns={
                    col_category: GL_COL_CATEGORY,
                    col_actuals: GL_COL_ACTUALS,
                    col_period: GL_COL_PERIOD,
                    col_activity: GL_COL_ACTIVITY,
                    col_fin_dept: GL_COL_FIN_DEPT,
                    col_desc: GL_COL_DESC,
                }
            )

            # Ignore N/A categories (and blanks)
            df_subset[GL_COL_CATEGORY] = df_subset[GL_COL_CATEGORY].apply(safe_str)
            df_subset = df_subset[
                df_subset[GL_COL_CATEGORY].str.upper().ne("N/A")
                & df_subset[GL_COL_CATEGORY].ne("")
            ]

            # Add Notes column for transactions
            df_subset["Notes"] = ""

            # Determine title department value
            dept_vals = [d for d in df_subset[GL_COL_FIN_DEPT].apply(safe_str).unique().tolist() if safe_str(d)]
            if len(dept_vals) == 1:
                dept_name = dept_vals[0]
            elif len(dept_vals) == 0:
                dept_name = "Financial Department"
            else:
                dept_name = "Multiple Departments"

            st.markdown("### GL Preview (transactions subset)")
            st.dataframe(df_subset.head(50), use_container_width=True)

            # Pivot: sum Actuals by Category
            df_subset[GL_COL_ACTUALS] = pd.to_numeric(df_subset[GL_COL_ACTUALS], errors="coerce").fillna(0.0)

            pivot = (
                df_subset.groupby(GL_COL_CATEGORY, dropna=False, as_index=False)[GL_COL_ACTUALS]
                .sum()
                .sort_values(by=GL_COL_ACTUALS, ascending=False, na_position="last")
            )

            st.markdown("### Pivot Preview (Actuals by Category)")
            st.dataframe(pivot, use_container_width=True)

            report_label = st.text_input("Report label (used in filename)", value=f"{dept_name} Expenses")

            if st.button("Generate GL Excel (transactions + pivot)", type="primary"):
                xlsx_bytes = build_gl_excel(
                    df_subset=df_subset,
                    pivot_categories_actuals=pivot[[GL_COL_CATEGORY, GL_COL_ACTUALS]].copy(),
                    dept_name=dept_name,
                )
                st.success("GL Excel generated!")
                st.download_button(
                    "Download GL Excel",
                    data=xlsx_bytes,
                    file_name=f"{make_safe_filename_fragment(report_label)}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            st.stop()

        # ============================================================
        # PPM MODE
        # ============================================================
        df_master, report_date = read_aggy_master(db_bytes)

        if report_date:
            st.info(f"PPM download date: **{report_date}**")
            date_label = report_date
        else:
            st.warning("Could not detect 'Report Run Date', please enter a date manually.")
            date_label = st.text_input("Report Date (YYYY-MM-DD)", value="")

        do_active_only = st.checkbox("Keep only ACTIVE Projects", value=True)
        status_col = find_column_by_exact_or_keywords(df_master.columns, STATUS_COL_NAME, keywords=["project", "status"])
        df_master_view = df_master[df_master[status_col] == "ACTIVE"].copy() if do_active_only else df_master.copy()

        df_award = None
        has_award = award_file is not None

        st.markdown("### Preview (before merge)")
        if has_award:
            award_bytes = award_file.getvalue()
            xl_aw = pd.ExcelFile(BytesIO(award_bytes))
            award_sheet = st.selectbox("Award sheet to use", options=xl_aw.sheet_names, index=0)
            df_award = read_award(award_bytes, sheet_name=award_sheet)

            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**PPM preview**")
                st.dataframe(df_master_view.head(25), use_container_width=True)
            with c2:
                st.markdown("**Award Info preview**")
                st.dataframe(df_award.head(25), use_container_width=True)

            st.markdown("---")
            st.markdown("### Choose which columns to merge by:")

            default_master_merge = PROJECT_COL_NAME if PROJECT_COL_NAME in df_master_view.columns else df_master_view.columns[0]

            default_aw_merge = None
            for cand in ["Aggie Enterprise Project #", "AGGIE ENTERPRISE PROJECT #", "AGGIE ENTERPRISE PROJECT # "]:
                if cand in df_award.columns:
                    default_aw_merge = cand
                    break
            if default_aw_merge is None:
                default_aw_merge = df_award.columns[0]

            master_merge_col = st.selectbox(
                "PPM merge column",
                options=list(df_master_view.columns),
                index=list(df_master_view.columns).index(default_master_merge) if default_master_merge in df_master_view.columns else 0,
            )
            award_merge_col = st.selectbox(
                "Award Document merge column",
                options=list(df_award.columns),
                index=list(df_award.columns).index(default_aw_merge) if default_aw_merge in df_award.columns else 0,
            )

            default_aw_rate = None
            for cand in ["INDIRECT RATE", "Indirect Rate", "Indirect rate"]:
                if cand in df_award.columns:
                    default_aw_rate = cand
                    break
            if default_aw_rate is None:
                indirect_candidates = [c for c in df_award.columns if "indirect" in c.lower()]
                default_aw_rate = indirect_candidates[0] if indirect_candidates else df_award.columns[-1]

            award_rate_col = st.selectbox(
                "Award Document indirect-rate column",
                options=list(df_award.columns),
                index=list(df_award.columns).index(default_aw_rate) if default_aw_rate in df_award.columns else 0,
            )

            master_keys = df_master_view[master_merge_col].apply(canon_key)
            award_keys = df_award[award_merge_col].apply(canon_key)

            master_key_set = set(k for k in master_keys.unique() if k)
            award_key_set = set(k for k in award_keys.unique() if k)
            intersect = master_key_set.intersection(award_key_set)
            match_rate_unique = (len(intersect) / len(master_key_set)) if master_key_set else 0.0

            st.markdown("### Merge preview")
            st.write(f"**# of PPM Projects:** {len(master_key_set)}")
            st.write(f"**# of Award Info Sheet Projects:** {len(award_key_set)}")
            st.write(f"**# that match:** {len(intersect)}")
            st.write(f"**Approx. match rate (unique PPM keys found in award):** {match_rate_unique:.1%}")

            if show_key_samples:
                st.markdown("**Project # Samples:**")
                c3, c4 = st.columns(2)
                with c3:
                    st.caption("PPM key sample")
                    st.code(", ".join(list(master_key_set)[:20]) if master_key_set else "(none)")
                with c4:
                    st.caption("Award key sample")
                    st.code(", ".join(list(award_key_set)[:20]) if award_key_set else "(none)")

            if show_unmatched:
                missing = sorted(list(master_key_set - award_key_set))[:40]
                if missing:
                    st.warning("Some PPM projects were not found in the Award document:")
                    st.code(", ".join(missing[:40]))
        else:
            st.dataframe(df_master_view.head(25), use_container_width=True)
            st.info("No Award Info document uploaded — reports will be generated WITHOUT indirect calculations.")

        st.markdown("---")
        st.markdown("### Generate Monthly Reports")

        if st.button("Generate ZIP (one Excel per PI)", type="primary"):
            df_work_full = df_master_view.copy()

            project_col = find_column_by_exact_or_keywords(df_work_full.columns, PROJECT_COL_NAME, keywords=["project", "number"])
            task_name_col = find_column_by_exact_or_keywords(df_work_full.columns, TASK_NAME_COL_NAME, keywords=["task", "name"])
            task_num_col = find_column_by_exact_or_keywords(df_work_full.columns, TASK_NUMBER_COL_NAME, keywords=["task", "number"])

            balance_candidates = [c for c in df_work_full.columns if str(c).startswith("Budget Balance")]
            if not balance_candidates:
                raise KeyError("PPM document is missing a column starting with 'Budget Balance'.")
            balance_col = balance_candidates[0]

            keep_cols = []
            for c in [
                PI_COL_NAME,
                project_col,
                "Project Name",
                "Project Manager",
                task_name_col,
                task_num_col,
                "Budget",
                "expenses",
                balance_col,
            ]:
                if c in df_work_full.columns and c not in keep_cols:
                    keep_cols.append(c)

            df_work = df_work_full[keep_cols].copy()

            p = df_work[project_col].apply(safe_str).str.replace(".0", "", regex=False)
            t = df_work[task_num_col].apply(safe_str).str.replace(".0", "", regex=False)
            df_work[project_col] = (p + "-" + t).str.strip("-")

            if "Project Name" in df_work.columns:
                df_work["Project Name"] = df_work["Project Name"].apply(safe_str) + " – " + df_work[task_name_col].apply(safe_str)

            df_work = df_work.drop(columns=[task_name_col, task_num_col], errors="ignore")
            df_work = df_work.rename(columns={"Budget": "Allocated Budget", balance_col: "Current Balance"})

            if has_award and df_award is not None:
                df_work["_merge_key"] = df_work_full[master_merge_col].apply(canon_key)

                df_aw = df_award.copy()
                df_aw["_merge_key"] = df_aw[award_merge_col].apply(canon_key)

                df_aw_sub = df_aw[["_merge_key", award_rate_col]].copy()
                df_aw_sub = df_aw_sub.drop_duplicates(subset=["_merge_key"], keep="first")
                df_aw_sub = df_aw_sub.rename(columns={award_rate_col: "Indirect Rate"})

                df_merged = df_work.merge(df_aw_sub, on="_merge_key", how="left").drop(columns=["_merge_key"])

                df_merged["Indirect Rate"] = pd.to_numeric(df_merged["Indirect Rate"], errors="coerce").fillna(0.0)
                df_merged["Allocated Budget"] = pd.to_numeric(df_merged["Allocated Budget"], errors="coerce")
                df_merged["Current Balance"] = pd.to_numeric(df_merged["Current Balance"], errors="coerce")

                denom = 1.0 + df_merged["Indirect Rate"]
                df_merged[ALLOC_BUDGET_NET_COL] = df_merged["Allocated Budget"] / denom
                df_merged[CURRENT_BAL_NET_COL] = df_merged["Current Balance"] / denom

                df_out = df_merged.drop(columns=["Allocated Budget", "Current Balance"], errors="ignore")
            else:
                df_work["Allocated Budget"] = pd.to_numeric(df_work["Allocated Budget"], errors="coerce")
                df_work["Current Balance"] = pd.to_numeric(df_work["Current Balance"], errors="coerce")
                df_work[ALLOC_BUDGET_NET_COL] = df_work["Allocated Budget"]
                df_work[CURRENT_BAL_NET_COL] = df_work["Current Balance"]
                df_out = df_work.drop(columns=["Allocated Budget", "Current Balance"], errors="ignore")

            if safe_str(date_label):
                df_out["Date Pulled"] = safe_str(date_label)

            if PI_COL_NAME not in df_out.columns:
                raise KeyError(f"PI column '{PI_COL_NAME}' not found in PPM. Columns: {list(df_out.columns)}")
            df_out[PI_COL_NAME] = df_out[PI_COL_NAME].apply(normalize_pi_last_first)

            desired = [
                PI_COL_NAME,
                "Project Manager",
                "Date Pulled",
                project_col,
                "Project Name",
                ALLOC_BUDGET_NET_COL,
                CURRENT_BAL_NET_COL,
                "Indirect Rate",
                "expenses",
            ]
            desired = [c for c in desired if c in df_out.columns]
            remaining = [c for c in df_out.columns if c not in desired]
            df_out = df_out[desired + remaining]

            report_label = "Budget Report"
            if safe_str(date_label):
                report_label = f"{safe_str(date_label)} Budget Report"

            hide_indirect_effective = hide_indirect_in_output if ("Indirect Rate" in df_out.columns) else True

            zip_bytes = build_pi_zip(
                df_out=df_out,
                pi_col=PI_COL_NAME,
                hide_indirect=hide_indirect_effective,
                report_label=report_label,
            )

            st.success("ZIP generated!")
            st.download_button(
                "Download ZIP (PI files)",
                data=zip_bytes,
                file_name=f"{make_safe_filename_fragment(report_label)} - PI Files.zip",
                mime="application/zip",
            )

    except Exception as e:
        st.error(f"Error: {e}")
        if show_trace:
            st.code(traceback.format_exc())
else:
    st.info("Choose PPM or GL, then upload the corresponding database file. (Award Info is only available in PPM mode.)")
